import time
import os
import openpyxl

current_folder = os.getcwd()
new_sku_list = []
counter = 1

os.system(f'python {current_folder}/Parse_OZON.py')
time.sleep(4)
os.system(f'python {current_folder}/CSV_to_XLSX.py')
time.sleep(4)
os.remove(f'{current_folder}/Ozon_Data.csv')

wb = openpyxl.load_workbook(f'{current_folder}/Ozon_Data.xlsx')
ws = wb.active

for cell in ws['C']:
    new_sku_list.append(cell.value)

wb.close()
os.remove(f'{current_folder}/Ozon_Data.xlsx')

wb = openpyxl.load_workbook(f'{current_folder}/OZON_DB_ids.xlsx')
ws = wb.active

ws.delete_cols(1, 1)

for i in new_sku_list:
    ws[f'A{counter}'] = i
    counter += 1

wb.save(f'{current_folder}/OZON_DB_ids.xlsx')

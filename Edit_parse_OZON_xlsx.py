import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

current_folder = os.getcwd()

wb = openpyxl.load_workbook(f'{current_folder}/Ozon_Data.xlsx')
ws = wb.active

ws.delete_cols(7, 1)

ws.column_dimensions[get_column_letter(1)].width = 10
ws.column_dimensions[get_column_letter(2)].width = 15
ws.column_dimensions[get_column_letter(3)].width = 12
ws.column_dimensions[get_column_letter(4)].width = 40
ws.column_dimensions[get_column_letter(5)].width = 15
ws.column_dimensions[get_column_letter(6)].width = 15
ws.column_dimensions[get_column_letter(7)].width = 15
ws.column_dimensions[get_column_letter(8)].width = 7
for i in range(9, 33):
    ws.column_dimensions[get_column_letter(i)].width = 7

ws['AE1'] = ' '
ws["A1"].alignment = Alignment(horizontal='center')
ws["B1"].alignment = Alignment(horizontal='center')
ws["C1"].alignment = Alignment(horizontal='center')
ws["D1"].alignment = Alignment(horizontal='center')
ws["E1"].alignment = Alignment(horizontal='center')
ws["F1"].alignment = Alignment(horizontal='center')
for i in range(71, 91):
    ws[f"{chr(i)}1"].alignment = Alignment(horizontal='left')
    ws[f"{chr(i)}1"].alignment = Alignment(vertical='bottom')

ws["AA1"].alignment = Alignment(horizontal='left')
ws["AB1"].alignment = Alignment(horizontal='left')
ws["AC1"].alignment = Alignment(horizontal='left')

for i in range(65, 68):
    for f in range(2, 300):
        ws[f"{chr(i)}{f}"].alignment = Alignment(horizontal='center')

for i in range(69, 91):
    for f in range(2, 300):
        ws[f"{chr(i)}{f}"].alignment = Alignment(horizontal='center')

for i in range(65,68):
    for f in range(2, 300):
        ws[f"A{chr(i)}{f}"].alignment = Alignment(horizontal='center')

wb.save(f'{current_folder}/Ozon_Data.xlsx')
